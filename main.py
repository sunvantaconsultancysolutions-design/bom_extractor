"""
BOM Extractor – Vercel-ready FastAPI app (FIXED VERSION)
Entrypoint: main.py  (Vercel auto-detects FastAPI here)

Key improvements:
- Better error messages with debugging info
- Proper async handling
- CORS explicitly configured
- Robust API key handling
- Better logging for Vercel troubleshooting
"""

import base64
import io
import json
import os
import traceback
import sys
from pathlib import Path
from typing import Optional

import openpyxl
import anthropic
from fastapi import FastAPI, Header, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel

# ─── Debugging: Log startup info ──────────────────────────────────────────────
print(f"[STARTUP] Python {sys.version}", file=sys.stderr)
print(f"[STARTUP] FastAPI app initializing...", file=sys.stderr)
API_KEY_CONFIGURED = bool(os.getenv("ANTHROPIC_API_KEY"))
print(f"[STARTUP] ANTHROPIC_API_KEY set: {API_KEY_CONFIGURED}", file=sys.stderr)

# ─── App ──────────────────────────────────────────────────────────────────────
app = FastAPI(title="BOM Extractor", version="2.0.1")

# ─── CORS with explicit origin handling ────────────────────────────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

# ─── Favicon (avoid 404) ──────────────────────────────────────────────────────
_FAVICON_ICO = base64.b64decode(
    "AAABAAEAAQEAAAEAGAAwAAAAFgAAACgAAAABAAAAAgAAAAEAGAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP8AAAA="
)

@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return Response(content=_FAVICON_ICO, media_type="image/x-icon")

# ─── Health check endpoint ─────────────────────────────────────────────────────
@app.get("/api/health", include_in_schema=False)
def health():
    return {
        "status": "ok",
        "api_key_configured": API_KEY_CONFIGURED,
        "version": "2.0.1"
    }

# ─── Serve frontend ───────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse, include_in_schema=False)
def root():
    """Serve the single-page frontend."""
    try:
        html_path = Path(__file__).parent / "index.html"
        if not html_path.exists():
            return "<h1>Error: index.html not found</h1>"
        return html_path.read_text(encoding='utf-8')
    except Exception as e:
        print(f"[ERROR] Failed to serve index.html: {e}", file=sys.stderr)
        return f"<h1>Error: {e}</h1>"

# ─── Claude system prompt ─────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert engineering drawing parser specialising in steel tower Bills of Materials (BOM).

Given a page image from an engineering drawing, extract ALL BOM tables visible on the page.

Return ONLY a valid JSON object (no markdown, no explanation) with this structure:
{
  "sections": [
    {
      "title": "BILL OF MATERIALS FOR (CAGE-1)",
      "unit": "QTY. PER (TOWER)",
      "profiles": [
        {"mark": "100H", "qty": 2, "profile": "L 90x6 H", "length": 4981, "black_weight": 82.66}
      ],
      "fasteners": [
        {"description": "M 16x40", "qty": 184, "black_weight": 28.46}
      ],
      "total_profiles_weight": 5100.80,
      "total_fasteners_weight": 210.83,
      "drawing_total_weight": 5311.62
    }
  ],
  "has_bom": true
}

If no BOM table is found on the page return: {"has_bom": false, "sections": []}

Rules:
- Extract every row — do not skip any.
- length is in mm (integer), black_weight is in kg (float).
- If a cell is empty/unclear use null.
- For fastener rows that have no separate length column, set length to null.
- Preserve the exact mark/description text (e.g. "207H", "150AH", "Filler Washer M 16 #8").
"""


# ─── Helpers ──────────────────────────────────────────────────────────────────
def extract_bom_from_image(client: anthropic.Anthropic, page_num: int, image_b64: str, media_type: str = "image/jpeg") -> dict:
    """Extract BOM from a single page image using Claude."""
    try:
        print(f"[API] Extracting BOM from page {page_num} (image size: {len(image_b64) / 1024:.1f} KB)", file=sys.stderr)
        
        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            system=SYSTEM_PROMPT,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}},
                    {"type": "text", "text": f"Extract all BOM tables from this engineering drawing page (page {page_num})."},
                ],
            }],
        )
        
        raw = msg.content[0].text.strip()
        print(f"[API] Claude response length: {len(raw)} chars", file=sys.stderr)
        
        # Handle markdown code blocks if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:].lstrip()
        
        result = json.loads(raw)
        print(f"[API] Successfully parsed JSON for page {page_num}", file=sys.stderr)
        return result
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON decode failed for page {page_num}: {e}", file=sys.stderr)
        print(f"[DEBUG] Raw response: {raw[:500]}", file=sys.stderr)
        raise
    except anthropic.APIError as e:
        print(f"[ERROR] Anthropic API error for page {page_num}: {e}", file=sys.stderr)
        raise


# ─── Excel builder ────────────────────────────────────────────────────────────
def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick():
    s = Side(style="medium", color="2E75B6")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
SUB_FILL  = PatternFill("solid", start_color="2E75B6")
SEC_FILL  = PatternFill("solid", start_color="D6E4F0")
ALT_FILL  = PatternFill("solid", start_color="EBF3FB")
TOT_FILL  = PatternFill("solid", start_color="FFF2CC")
GTOT_FILL = PatternFill("solid", start_color="F4B942")
WHT_FILL  = PatternFill("solid", start_color="FFFFFF")
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
R = Alignment(horizontal="right",  vertical="center")


def sc(cell, val, fill=None, font=None, align=None, fmt=None):
    cell.value = val
    cell.fill  = fill or WHT_FILL
    cell.font  = font or Font(name="Arial", size=9)
    cell.alignment = align or L
    cell.border = _thin()
    if fmt:
        cell.number_format = fmt


def build_excel(all_sections: list[dict], filename: str) -> bytes:
    """Build Excel workbook from BOM sections."""
    try:
        print(f"[EXCEL] Building workbook with {len(all_sections)} sections", file=sys.stderr)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        summary_rows = []

        for sec in all_sections:
            title_raw = sec.get("title", "BOM")
            short = title_raw.replace("BILL OF MATERIALS FOR", "").strip("() ").replace("/", "-")[:31] or "BOM"
            ws = wb.create_sheet(title=short)
            ws.sheet_view.showGridLines = False

            ws.merge_cells("A1:F1")
            c = ws["A1"]; c.value = title_raw
            c.fill = HDR_FILL; c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.alignment = C; c.border = _thin(); ws.row_dimensions[1].height = 22

            ws.merge_cells("A2:F2")
            c = ws["A2"]; c.value = sec.get("unit", "")
            c.fill = SUB_FILL; c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            c.alignment = C; c.border = _thin(); ws.row_dimensions[2].height = 14

            row = 3
            for col, h in enumerate(["MARK","QTY","PROFILE","LENGTH (mm)","BLACK WEIGHT (kg)","NOTES"], 1):
                c = ws.cell(row=row, column=col); c.value = h; c.fill = SUB_FILL
                c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
                c.alignment = C; c.border = _thin()
            ws.row_dimensions[row].height = 26; row += 1

            profiles = sec.get("profiles", [])
            if profiles:
                ws.merge_cells(f"A{row}:F{row}")
                c = ws.cell(row=row, column=1); c.value = "▶  STRUCTURAL PROFILES"
                c.fill = SEC_FILL; c.font = Font(name="Arial", bold=True, size=9, color="1F4E79")
                c.alignment = L; c.border = _thin(); ws.row_dimensions[row].height = 14; row += 1
                for i, p in enumerate(profiles):
                    fill = ALT_FILL if i % 2 == 0 else WHT_FILL
                    ws.row_dimensions[row].height = 13
                    sc(ws.cell(row=row,column=1), p.get("mark"),         fill, fmt="@")
                    sc(ws.cell(row=row,column=2), p.get("qty"),          fill, align=C)
                    sc(ws.cell(row=row,column=3), p.get("profile"),      fill)
                    sc(ws.cell(row=row,column=4), p.get("length"),       fill, align=R, fmt="#,##0")
                    sc(ws.cell(row=row,column=5), p.get("black_weight"), fill, align=R, fmt="#,##0.00")
                    sc(ws.cell(row=row,column=6), "",                    fill); row += 1

            tp = sec.get("total_profiles_weight")
            if tp is not None:
                ws.merge_cells(f"A{row}:D{row}")
                c = ws.cell(row=row, column=1); c.value = "TOTAL BLACK WEIGHT (profiles) kg"
                c.fill = TOT_FILL; c.font = Font(name="Arial", bold=True, size=9, color="7F3F00")
                c.alignment = L; c.border = _thin()
                sc(ws.cell(row=row,column=5), tp, TOT_FILL, Font(name="Arial",bold=True,size=9,color="7F3F00"), R, "#,##0.00")
                sc(ws.cell(row=row,column=6), "", TOT_FILL); ws.row_dimensions[row].height = 15; row += 1

            fasteners = sec.get("fasteners", [])
            if fasteners:
                ws.merge_cells(f"A{row}:F{row}")
                c = ws.cell(row=row, column=1); c.value = "▶  FASTENERS"
                c.fill = SEC_FILL; c.font = Font(name="Arial", bold=True, size=9, color="1F4E79")
                c.alignment = L; c.border = _thin(); ws.row_dimensions[row].height = 14; row += 1
                for col, h in enumerate(["DESCRIPTION","QTY","BLACK WEIGHT (kg)","","",""], 1):
                    c = ws.cell(row=row, column=col); c.value = h; c.fill = SUB_FILL
                    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
                    c.alignment = C; c.border = _thin()
                ws.row_dimensions[row].height = 14; row += 1
                for i, f in enumerate(fasteners):
                    fill = ALT_FILL if i % 2 == 0 else WHT_FILL
                    ws.merge_cells(f"A{row}:B{row}")
                    sc(ws.cell(row=row,column=1), f.get("description"), fill)
                    sc(ws.cell(row=row,column=2), "",                   fill)
                    sc(ws.cell(row=row,column=3), f.get("qty"),         fill, align=C)
                    sc(ws.cell(row=row,column=4), f.get("black_weight"),fill, align=R, fmt="#,##0.00")
                    sc(ws.cell(row=row,column=5), "", fill); sc(ws.cell(row=row,column=6), "", fill)
                    ws.row_dimensions[row].height = 13; row += 1

                tf = sec.get("total_fasteners_weight")
                if tf is not None:
                    ws.merge_cells(f"A{row}:C{row}")
                    c = ws.cell(row=row, column=1); c.value = "TOTAL BLACK WEIGHT (fasteners) kg"
                    c.fill = TOT_FILL; c.font = Font(name="Arial", bold=True, size=9, color="7F3F00")
                    c.alignment = L; c.border = _thin()
                    sc(ws.cell(row=row,column=4), tf, TOT_FILL, Font(name="Arial",bold=True,size=9,color="7F3F00"), R, "#,##0.00")
                    for col in [2,5,6]: sc(ws.cell(row=row,column=col), "", TOT_FILL)
                    ws.row_dimensions[row].height = 15; row += 1

            dt = sec.get("drawing_total_weight")
            if dt is not None:
                ws.merge_cells(f"A{row}:D{row}")
                c = ws.cell(row=row, column=1); c.value = "DRAWING TOTAL BLACK WEIGHT (kg)"
                c.fill = GTOT_FILL; c.font = Font(name="Arial", bold=True, size=10)
                c.alignment = L; c.border = _thick()
                c2 = ws.cell(row=row,column=5); c2.value = dt; c2.fill = GTOT_FILL
                c2.font = Font(name="Arial",bold=True,size=10); c2.alignment = R
                c2.border = _thick(); c2.number_format = "#,##0.00"
                sc(ws.cell(row=row,column=6), "", GTOT_FILL); ws.row_dimensions[row].height = 18

            ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 6
            ws.column_dimensions["C"].width = 18; ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 20; ws.column_dimensions["F"].width = 10
            ws.freeze_panes = "A4"

            summary_rows.append({"section": short, "title": title_raw,
                                  "profiles": tp, "fasteners": sec.get("total_fasteners_weight"), "total": dt})

        # Summary sheet
        ws = wb.create_sheet("SUMMARY", 0)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 22; ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 22

        ws.merge_cells("A1:E1"); c = ws["A1"]; c.value = "BILL OF MATERIALS – SUMMARY"
        c.fill = HDR_FILL; c.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
        c.alignment = C; c.border = _thin(); ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:E2"); c = ws["A2"]; c.value = f"Source: {filename}"
        c.fill = SUB_FILL; c.font = Font(name="Arial", color="FFFFFF", size=9)
        c.alignment = C; c.border = _thin(); ws.row_dimensions[2].height = 14

        row = 4
        for col, h in enumerate(["SHEET","COMPONENT / SECTION","PROFILES WT. (kg)","FASTENERS WT. (kg)","DRAWING TOTAL (kg)"], 1):
            c = ws.cell(row=row, column=col); c.value = h; c.fill = SUB_FILL
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            c.alignment = C; c.border = _thin()
        ws.row_dimensions[row].height = 26; row += 1

        gp = gf = gt = 0.0
        for i, sr in enumerate(summary_rows):
            fill = ALT_FILL if i % 2 == 0 else WHT_FILL
            ws.row_dimensions[row].height = 15
            sc(ws.cell(row=row,column=1), sr["section"],   fill, Font(name="Arial",bold=True,size=9), C)
            sc(ws.cell(row=row,column=2), sr["title"],     fill)
            sc(ws.cell(row=row,column=3), sr["profiles"],  fill, align=R, fmt="#,##0.00")
            sc(ws.cell(row=row,column=4), sr["fasteners"], fill, align=R, fmt="#,##0.00")
            sc(ws.cell(row=row,column=5), sr["total"],     fill, Font(name="Arial",bold=True,size=9), R, "#,##0.00")
            gp += sr["profiles"] or 0; gf += sr["fasteners"] or 0; gt += sr["total"] or 0; row += 1

        ws.merge_cells(f"A{row}:B{row}"); c = ws.cell(row=row, column=1); c.value = "GRAND TOTAL"
        c.fill = GTOT_FILL; c.font = Font(name="Arial", bold=True, size=11)
        c.alignment = C; c.border = _thick()
        for col, val in [(3, round(gp,2)),(4, round(gf,2)),(5, round(gt,2))]:
            c = ws.cell(row=row, column=col); c.value = val; c.fill = GTOT_FILL
            c.font = Font(name="Arial",bold=True,size=11); c.alignment = R
            c.border = _thick(); c.number_format = "#,##0.00"
        ws.row_dimensions[row].height = 22; ws.freeze_panes = "A5"

        buf = io.BytesIO(); wb.save(buf); 
        print(f"[EXCEL] Workbook created successfully ({len(buf.getvalue()) / 1024:.1f} KB)", file=sys.stderr)
        return buf.getvalue()
    except Exception as e:
        print(f"[ERROR] Excel build failed: {e}", file=sys.stderr)
        print(f"[TRACEBACK] {traceback.format_exc()}", file=sys.stderr)
        raise


# ─── API routes ───────────────────────────────────────────────────────────────
class PageExtractRequest(BaseModel):
    """Single page image sent as base64 for BOM extraction."""
    page_num: int
    image_data: str      # base64-encoded JPEG/PNG
    media_type: str = "image/jpeg"


@app.post("/api/extract-page")
async def extract_page(
    payload: PageExtractRequest,
    x_api_key: Optional[str] = Header(default=None, alias="X-API-Key"),
):
    """Extract BOM from a single page image. Called once per page from the browser."""
    api_key = x_api_key or os.getenv("ANTHROPIC_API_KEY", "").strip()
    
    if not api_key:
        print(f"[AUTH] No API key provided (header={bool(x_api_key)}, env={bool(os.getenv('ANTHROPIC_API_KEY'))})", file=sys.stderr)
        raise HTTPException(
            status_code=401,
            detail="Anthropic API key required. Set ANTHROPIC_API_KEY environment variable or provide X-API-Key header."
        )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        result = extract_bom_from_image(client, payload.page_num, payload.image_data, payload.media_type)
        return result
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON decode error on page {payload.page_num}: {e}", file=sys.stderr)
        raise HTTPException(
            status_code=422,
            detail=f"Claude returned invalid JSON. This may indicate the PDF page was unreadable. Error: {str(e)[:100]}"
        )
    except anthropic.APIError as e:
        error_msg = str(e)
        print(f"[ERROR] Anthropic API error on page {payload.page_num}: {error_msg}", file=sys.stderr)
        # Check for specific error types
        if "invalid_api_key" in error_msg.lower():
            raise HTTPException(status_code=401, detail="Invalid Anthropic API key. Check your environment variables.")
        elif "rate_limit" in error_msg.lower():
            raise HTTPException(status_code=429, detail="Rate limited by Anthropic API. Please wait a moment and try again.")
        raise HTTPException(status_code=502, detail=f"Anthropic API error: {error_msg[:100]}")
    except Exception as e:
        print(f"[ERROR] Unexpected error on page {payload.page_num}: {e}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)[:100]}")


class BuildExcelRequest(BaseModel):
    """All collected BOM sections + filename to build the Excel workbook."""
    filename: str
    sections: list[dict]


@app.post("/api/build-excel")
async def build_excel_endpoint(payload: BuildExcelRequest):
    """Build Excel from collected BOM sections. Called once after all pages are processed."""
    if not payload.sections:
        raise HTTPException(status_code=422, detail="No BOM sections provided.")

    try:
        xlsx_bytes = build_excel(payload.sections, payload.filename)
        stem = Path(payload.filename).stem
        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}_BOM.xlsx"'},
        )
    except Exception as e:
        print(f"[ERROR] Excel build endpoint failed: {e}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Excel build failed: {str(e)[:100]}")